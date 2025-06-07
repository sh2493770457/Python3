import argparse
import os
import re
import requests
import json
import sys
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# 控制台颜色配置
COLOR = {
    "HEADER": "\033[95m",
    "OKBLUE": "\033[94m",
    "OKCYAN": "\033[96m",
    "OKGREEN": "\033[92m",
    "WARNING": "\033[93m",
    "FAIL": "\033[91m",
    "ENDC": "\033[0m",
    "BOLD": "\033[1m",
    "UNDERLINE": "\033[4m"
}

class RateLimiter:
    def __init__(self):
        self.remaining = 10
        self.reset_time = time.time() + 3600
        self.last_request = 0

    def update(self, headers):
        if 'X-RateLimit-Remaining' in headers:
            self.remaining = int(headers.get('X-RateLimit-Remaining', 10))
            self.reset_time = int(headers.get('X-RateLimit-Reset', time.time() + 3600))

    def wait(self):
        now = time.time()
        min_interval = 1.1
        if now - self.last_request < min_interval:
            time.sleep(min_interval - (now - self.last_request))
        if self.remaining <= 3:
            sleep_time = max(self.reset_time - now + 2, 15)
            print(f"{COLOR['WARNING']}⚠ 预防性等待 {int(sleep_time)}秒 (剩余请求: {self.remaining}){COLOR['ENDC']}")
            time.sleep(sleep_time)
        self.last_request = time.time()

def check_github_token():
    token = os.getenv("GITHUB_TOKEN")
    if not token:
        print(f"{COLOR['FAIL']}✘ 必须配置GitHub访问令牌！{COLOR['ENDC']}")
        print("Windows正确配置方法：")
        print('   set GITHUB_TOKEN=你的令牌值')
        sys.exit(1)
    return token

def parse_args():
    parser = argparse.ArgumentParser(description='GitHub高级项目检索工具')
    parser.add_argument('--word', type=str, help='搜索关键词')
    parser.add_argument('--output', type=str, default='md', choices=['json', 'xlsx', 'md', 'txt'])
    parser.add_argument('--min-stars', type=int, default=0)
    parser.add_argument('--min-forks', type=int, default=0)
    parser.add_argument('--updated-after', type=str)
    parser.add_argument('--sort-by', type=str, default='stars', choices=['stars', 'forks', 'updated'])
    parser.add_argument('--order', type=str, default='desc', choices=['desc', 'asc'])
    parser.add_argument('--has-readme', action='store_true')
    parser.add_argument('--limit', type=int, default=100)
    parser.add_argument('--top', type=int, default=0)
    
    args = parser.parse_args()
    
    if len(sys.argv) > 1:
        if args.top == 0 and not args.word:
            parser.error("必须使用 --word 或 --top 参数")
        if args.top > 0:
            if args.word or args.min_stars > 0 or args.min_forks > 0:
                parser.error("--top 模式不能与其他参数共用")
            if args.top > 1000:
                args.top = 1000
                print(f"{COLOR['WARNING']}警告：结果数已自动限制为1000{COLOR['ENDC']}")
        else:
            if args.limit > 1000:
                args.limit = 1000
                print(f"{COLOR['WARNING']}警告：结果数已自动限制为1000{COLOR['ENDC']}")
    return args

def get_interactive_input():
    print(f"\n{COLOR['HEADER']}=== GitHub项目检索工具（交互模式）==={COLOR['ENDC']}")
    
    def validate(prompt, default, check, convert):
        while True:
            try:
                val = input(f"{COLOR['OKCYAN']}{prompt}（默认：{default}）: {COLOR['ENDC']}").strip()
                val = val or str(default)
                if not check(val):
                    raise ValueError
                return convert(val)
            except:
                print(f"{COLOR['FAIL']}输入无效，请重试{COLOR['ENDC']}")
    
    word = validate("搜索关键词", "", lambda x: len(x)>=2, str)
    if not word:
        print(f"{COLOR['FAIL']}✘ 必须输入搜索关键词{COLOR['ENDC']}")
        sys.exit(1)
        
    return argparse.Namespace(
        word=word,
        min_stars=validate("最小Star数", 0, lambda x: x.isdigit(), int),
        min_forks=validate("最小Fork数", 0, lambda x: x.isdigit(), int),
        updated_after=validate("更新时间（YYYY-MM-DD）", "", lambda x: True, str),
        has_readme=validate("必须包含README（y/n）", "n", lambda x: x.lower() in ['y','n'], lambda x: x == 'y'),
        sort_by=validate("排序字段", "stars", lambda x: x in ['stars','forks','updated'], str),
        order=validate("排序顺序", "desc", lambda x: x in ['desc','asc'], str),
        output=validate("导出格式", "md", lambda x: x in ['json','xlsx','md','txt'], str),
        limit=validate("结果数量", 100, lambda x: x.isdigit() and 1<=int(x)<=1000, lambda x: min(int(x),1000)),
        top=0
    )

def build_query(args):
    if args.top > 0:
        return "stars:>1"
    
    if not args.word:
        print(f"{COLOR['FAIL']}✘ 必须提供搜索关键词{COLOR['ENDC']}")
        sys.exit(1)
        
    query = [args.word]
    if args.min_stars > 0:
        query.append(f"stars:>={args.min_stars}")
    if args.min_forks > 0:
        query.append(f"forks:>={args.min_forks}")
    if args.updated_after:
        try:
            datetime.strptime(args.updated_after, '%Y-%m-%d')
            query.append(f"pushed:>={args.updated_after}")
        except ValueError:
            print(f"{COLOR['FAIL']}日期格式错误，已忽略该条件{COLOR['ENDC']}")
    if args.has_readme:
        query.append("has:readme")
    return ' '.join(query)

def search_github(args):
    token = check_github_token()
    rate_limiter = RateLimiter()
    headers = {'Authorization': f'token {token}'}
    session = requests.Session()
    
    print(f"\n{COLOR['OKGREEN']}▶ 正在搜索GitHub...{COLOR['ENDC']}")
    
    results = []
    page = 1
    target = args.top if args.top > 0 else args.limit
    
    while len(results) < target:
        rate_limiter.wait()
        
        try:
            params = {
                'q': build_query(args),
                'sort': 'stars' if args.top > 0 else args.sort_by,
                'order': 'desc' if args.top > 0 else args.order,
                'per_page': min(100, target - len(results)),
                'page': page
            }
            
            response = session.get('https://api.github.com/search/repositories',
                                 params=params, headers=headers, timeout=15)
            rate_limiter.update(response.headers)
            response.raise_for_status()
            
            items = response.json().get('items', [])
            results.extend(items)
            page += 1
            
            print(f"{COLOR['OKBLUE']}↳ 已获取 {len(results)}/{target} 剩余请求: {rate_limiter.remaining}{COLOR['ENDC']}", end='\r')
            
            if not items:
                break
                
        except requests.exceptions.RequestException as e:
            print(f"\n{COLOR['FAIL']}✘ 请求失败: {str(e)}{COLOR['ENDC']}")
            if hasattr(e, 'response') and e.response.status_code == 403:
                wait_time = max(rate_limiter.reset_time - time.time(), 15)
                print(f"{COLOR['WARNING']}→ 等待{int(wait_time)}秒后重试...{COLOR['ENDC']}")
                time.sleep(wait_time)
            else:
                raise
    
    print(f"\n{COLOR['OKGREEN']}✔ 成功获取 {len(results)} 条数据{COLOR['ENDC']}")
    return results[:target]

def process_results(items, args):  # 修改后的函数签名
    valid = []
    for item in items:
        desc = (item.get('description') or '')[:200]
        if re.search(r'[=\.\-_]{4,}', desc):
            continue
            
        # 新增相关性检查（仅当有搜索关键词时生效）
        if args.word:
            keywords = args.word.lower().split()
            name_lower = item['name'].lower()
            desc_lower = desc.lower()
            
            # 检查是否至少匹配一个关键词
            has_keyword = any(
                kw in name_lower or kw in desc_lower
                for kw in keywords
            )
            if not has_keyword:
                continue
                
        valid.append({
            'name': item['name'],
            'description': desc,
            'stars': item['stargazers_count'],
            'forks': item['forks_count'],
            'url': item['html_url'],
            'updated': item['pushed_at'][:10]
        })
    return valid

def export_to_json(results, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"{COLOR['OKGREEN']}JSON文件已生成：{filename}{COLOR['ENDC']}")

def export_to_xlsx(results, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "GitHub项目"
    
    headers = ["排名", "项目名称", "描述", "Stars", "Forks", "更新时间", "链接"]
    ws.append(headers)
    
    for col in 'ABCDEFG':
        ws.column_dimensions[col].width = 20
        
    for idx, item in enumerate(results, 1):
        row = [idx, item['name'], item['description'], item['stars'], item['forks'], item['updated']]
        ws.append(row)
        cell = ws[f'G{idx+1}']
        cell.value = '访问项目'
        cell.hyperlink = item['url']
        cell.font = Font(color="0000FF", underline="single")
    
    wb.save(filename)
    print(f"{COLOR['OKGREEN']}Excel文件已生成：{filename}{COLOR['ENDC']}")

def export_to_md(results, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        f.write("# GitHub项目检索结果\n\n")
        f.write("| 排名 | 项目名称 | 描述 | Stars | Forks | 更新时间 | 链接 |\n")
        f.write("|------|----------|------|-------|-------|----------|------|\n")
        for idx, item in enumerate(results, 1):
            f.write(f"| {idx} | [{item['name']}]({item['url']}) | {item['description']} | ⭐ {item['stars']} | 🍴 {item['forks']} | {item['updated']} | [访问]({item['url']}) |\n")
    print(f"{COLOR['OKGREEN']}Markdown文件已生成：{filename}{COLOR['ENDC']}")

def export_to_txt(results, filename):
    with open(filename, 'w', encoding='utf-8') as f:
        f.write("GitHub项目检索结果\n")
        f.write("="*80 + "\n")
        for idx, item in enumerate(results, 1):
            f.write(f"{idx}. {item['name']}\n")
            f.write(f"   描述: {item['description']}\n")
            f.write(f"   Stars: {item['stars']} | Forks: {item['forks']}\n")
            f.write(f"   更新: {item['updated']}\n")
            f.write(f"   链接: {item['url']}\n")
            f.write("-"*80 + "\n")
    print(f"{COLOR['OKGREEN']}文本文件已生成：{filename}{COLOR['ENDC']}")

if __name__ == '__main__':
    try:
        if len(sys.argv) == 1:
            args = get_interactive_input()
        else:
            args = parse_args()
        check_github_token()
        
        if args.top > 0:
            filename = f"top{args.top}_stars.{args.output}"
            args.sort_by = 'stars'
            args.order = 'desc'
            args.limit = args.top
        else:
            filename = f"{args.word}_results.{args.output}"
            
        raw_data = search_github(args)
        final_results = process_results(raw_data, args)  # 传递args参数
        
        if not final_results:
            print(f"{COLOR['WARNING']}⚠ 未找到有效结果{COLOR['ENDC']}")
            sys.exit()
            
        {
            'json': lambda: export_to_json(final_results, filename),
            'xlsx': lambda: export_to_xlsx(final_results, filename),
            'md': lambda: export_to_md(final_results, filename),
            'txt': lambda: export_to_txt(final_results, filename)
        }[args.output.lower()]()
        
    except Exception as e:
        print(f"\n{COLOR['FAIL']}✘ 运行错误：{str(e)}{COLOR['ENDC']}")
        sys.exit(1)